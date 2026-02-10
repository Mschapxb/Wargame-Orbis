"""Bibliothèque d'unités parsée depuis le fichier Excel Orbis Naturae.

Dépendance unique: openpyxl (pip install openpyxl)
Les traits spéciaux sont ignorés pour l'instant.
"""

import os
from openpyxl import load_workbook
from models import Arme
from unit import Unit


def _parse_int(val, default=0):
    if val is None:
        return default
    s = str(val).strip()
    if s in ('', 'nan', '—', '-', '—*', '_x007f_—', 'Intouchable', 'Fuir ? Naaaa'):
        return default
    s = s.replace('M', '').replace('m', '').replace('*', '').replace('²', '')
    if '/' in s:
        s = s.split('/')[0]
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default


def _parse_speed(val):
    if val is None:
        return 2
    s = str(val).strip()
    if s in ('**', '—*', '', 'nan', 'Trop rapide'):
        return 2
    s = s.replace('M', '').replace('m', '').replace('*', '').replace('/', '')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 2


def _parse_degats(val):
    if val is None:
        return "1"
    s = str(val).strip().upper()
    if s in ('', 'nan', '0', 'AÏE', '0²'):
        return "1"
    s = s.replace('*', '').replace('²', '')
    if '/' in s:
        s = s.split('/')[0]
    s = s.lower().strip()
    return s if s else "1"


def _parse_portee(val):
    if val is None:
        return 1
    s = str(val).strip()
    if s in ('', 'nan', '—²', 'Infini', 'None'):
        return 1
    s = s.replace('*', '').replace('²', '')
    if '-' in s:
        parts = s.split('-')
        try:
            return max(int(p) for p in parts)
        except ValueError:
            return 1
    try:
        raw = int(float(s))
        if raw > 10:
            return max(2, raw // 4)
        return max(1, raw)
    except (ValueError, TypeError):
        return 1


def _parse_attaque(val):
    if val is None:
        return 1
    s = str(val).strip()
    if s in ('', 'nan', '**', '—*', 'None'):
        return 1
    s = s.replace('*', '').replace('²', '')
    if '/' in s:
        s = s.split('/')[0]
    try:
        return max(1, int(float(s)))
    except (ValueError, TypeError):
        return 1


def _cell(row, col):
    """Récupère la valeur d'une cellule d'une ligne openpyxl."""
    if col < len(row):
        v = row[col]
        return v if v is not None else None
    return None


def _build_arme(arme_name, portee, attaque, toucher, blesser, perforation, degats):
    name = str(arme_name).strip() if arme_name else ''
    if not name or name == 'None':
        return None
    return Arme(name,
                nb_attaque=_parse_attaque(attaque),
                toucher=_parse_int(toucher, 3),
                blesser=_parse_int(blesser, 3),
                perforation=_parse_int(perforation, 0),
                degats=_parse_degats(degats),
                porte=_parse_portee(portee))


def _determine_role(unit_data):
    portee_main = _parse_portee(unit_data['portee'])
    if portee_main >= 5:
        return "back"
    elif portee_main >= 3:
        return "mid"
    return "front"


def _army_color(army_name):
    colors = {
        'Armée Skaldienne': (80, 140, 200),
        'Collège de magie': (120, 80, 200),
        'Ordre de Chevalerie': (200, 180, 60),
        'Ordre Eternel': (100, 100, 100),
        'Invocation Eternel': (80, 80, 80),
        'Légion sacrée': (220, 200, 60),
        'Héros': (255, 215, 0),
        'Armée Orlandar': (60, 160, 60),
        'Draconie': (200, 60, 60),
        'Arkkar': (180, 120, 60),
        'Al-Athar': (160, 80, 160),
        'Armée Aïdatienne': (60, 180, 180),
        'Armée Marcheurs Jaunes': (200, 200, 40),
        'Kretash le Bouffon': (180, 80, 80),
        'Armée Muhr': (140, 100, 60),
        'Armée Huǒ shé': (220, 80, 40),
        'Clan Marsik': (100, 140, 100),
        'Clan Firast': (200, 100, 40),
        'Invocations du Tao': (100, 180, 100),
        'Unité commune': (150, 150, 150),
    }
    return colors.get(army_name, (150, 150, 150))


def load_units_from_excel(filepath):
    """Parse le fichier Excel et retourne {army_name: [unit_data_dict, ...]}"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    all_units = {}
    current_army = None
    current_unit_data = None

    for raw_row in ws.iter_rows(values_only=True):
        # Padder la ligne pour éviter les IndexError
        row = list(raw_row) + [None] * 16
        val0 = str(row[0]).strip() if row[0] is not None else ''
        val1 = str(row[1]).strip() if row[1] is not None else ''

        if not val0 or val0 == 'None':
            # Ligne d'arme supplémentaire ?
            arme_name = str(row[5]).strip() if row[5] is not None else ''
            if arme_name and arme_name != 'None' and current_unit_data is not None:
                current_unit_data['extra_weapons'].append({
                    'arme': arme_name,
                    'portee': _cell(row, 6), 'attaque': _cell(row, 7),
                    'toucher': _cell(row, 8), 'blesser': _cell(row, 9),
                    'perforation': _cell(row, 10), 'degats': _cell(row, 11)
                })
            continue

        if val0 == 'Nom':
            continue

        # En-tête d'armée (pas de stats)
        if (not val1 or val1 == 'None') and row[2] is None:
            current_army = val0
            if current_army not in all_units:
                all_units[current_army] = []
            continue

        # Ligne d'unité
        if val1 and val1 != 'None' and current_army is not None:
            current_unit_data = {
                'army': current_army,
                'name': val0,
                'deplacement': val1,
                'blessure': _cell(row, 2),
                'bravoure': _cell(row, 3),
                'sauvegarde': _cell(row, 4),
                'arme': str(row[5]) if row[5] else '',
                'portee': _cell(row, 6),
                'attaque': _cell(row, 7),
                'toucher': _cell(row, 8),
                'blesser': _cell(row, 9),
                'perforation': _cell(row, 10),
                'degats': _cell(row, 11),
                'extra_weapons': []
            }
            all_units[current_army].append(current_unit_data)

    wb.close()
    return all_units


def create_unit_from_data(data):
    """Convertit un dict de données brutes en objet Unit."""
    armes = []
    main_arme = _build_arme(data['arme'], data['portee'], data['attaque'],
                            data['toucher'], data['blesser'],
                            data['perforation'], data['degats'])
    if main_arme:
        armes.append(main_arme)

    for w in data.get('extra_weapons', []):
        extra = _build_arme(w['arme'], w['portee'], w['attaque'],
                            w['toucher'], w['blesser'],
                            w['perforation'], w['degats'])
        if extra:
            armes.append(extra)

    unit = Unit(
        name=data['name'][:10],
        pv=_parse_int(data['blessure'], 1),
        vitesse=_parse_speed(data['deplacement']),
        morale=_parse_int(data['bravoure'], 1),
        sauvegarde=_parse_int(data['sauvegarde'], 7),
        color=_army_color(data.get('army', '')),
        armes=armes,
        role=_determine_role(data)
    )
    # Nom complet nettoyé pour le fichier token: "tokens/Infanterie régulière.png"
    unit.token_name = data['name'].strip()
    return unit


_DEFAULT_EXCEL = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'Livre_des_armées_d_Orbis_Naturae.xlsx')


def get_library(filepath=None):
    return load_units_from_excel(filepath or _DEFAULT_EXCEL)


def list_armies(filepath=None):
    return sorted(get_library(filepath).keys())


def list_units(army_name, filepath=None):
    return [u['name'] for u in get_library(filepath).get(army_name, [])]


def make_unit(army_name, unit_name, filepath=None):
    for u_data in get_library(filepath).get(army_name, []):
        if u_data['name'] == unit_name:
            return create_unit_from_data(u_data)
    return None
